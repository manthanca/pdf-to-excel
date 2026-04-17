#!/usr/bin/env python3
"""
MRACA Smart Contract Note Converter
PDF-to-Excel aggregator for Indian broker contract notes.
Run: streamlit run app_final.py
"""

import io
import os
import shutil
import sys
import time
import zipfile
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from universal_angel_one_processor import extract_from_angel_one_pdf, create_excel_output, parse_holding_statement_balances
from core.tax_engine import create_capital_gains_summary_sheet

# ---------------------------------------------------------------------------
# Page config
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="MRACA Smart Contract Note Converter",
    page_icon="💼",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ---------------------------------------------------------------------------
# Theme CSS
# ---------------------------------------------------------------------------
st.markdown("""
<style>
    .main-header {
        font-size: 2.2rem; font-weight: 700; text-align: center;
        background: linear-gradient(135deg, #1e3a5f 0%, #334155 100%);
        -webkit-background-clip: text; -webkit-text-fill-color: transparent;
        background-clip: text; margin-bottom: 0.3rem;
    }
    .sub-header {
        font-size: 1.1rem; color: #64748b; text-align: center;
        margin-bottom: 1.8rem; font-weight: 500;
    }
    .client-greeting {
        font-size: 1.8rem; font-weight: 600; color: #1e3a5f; margin-bottom: 1rem;
    }
    .success-message {
        background-color: #10b981; color: white; padding: 0.9rem;
        border-radius: 0.5rem; text-align: center; font-weight: 600;
    }
    .stProgress > div > div > div > div { background-color: #3b82f6; }
</style>
""", unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# Session-state initialisation
# ---------------------------------------------------------------------------
DEFAULTS = {
    "active_client": None,
    "total_pdfs_uploaded": 0,
    "processing_complete": False,
    "master_trades": [],
    "master_obligations": [],
    "show_client_creation": False,
    "show_client_selection": False,
    "master_excel_data": None,
    "master_excel_filename": "",
    "individual_files": [],
}

for key, val in DEFAULTS.items():
    if key not in st.session_state:
        st.session_state[key] = val


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

MONTH_MAP = {
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4,
    "May": 5, "Jun": 6, "Jul": 7, "Aug": 8,
    "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
}

QUARTER_MAP = {
    "Q1 (Apr-Jun)": (4, 6),
    "Q2 (Jul-Sep)": (7, 9),
    "Q3 (Oct-Dec)": (10, 12),
    "Q4 (Jan-Mar)": (1, 3),
}


def apply_date_filter(trades: list, date_filter: str, **kwargs) -> list:
    """Filter trade records by the selected date period."""
    if date_filter == "All Data" or not trades:
        return trades

    df = pd.DataFrame(trades)
    df["Trade Date"] = pd.to_datetime(df["Trade Date"], errors="coerce")

    if date_filter == "Quarterly (Q1-Q4)":
        quarter = kwargs.get("quarter", "Q1 (Apr-Jun)")
        start_m, end_m = QUARTER_MAP.get(quarter, (4, 6))
        if start_m <= end_m:
            df = df[df["Trade Date"].dt.month.between(start_m, end_m)]
        else:  # Q4 wraps year (Jan-Mar)
            df = df[df["Trade Date"].dt.month.isin(range(start_m, 13)) |
                    df["Trade Date"].dt.month.isin(range(1, end_m + 1))]

    elif date_filter == "Monthly":
        month_num = MONTH_MAP.get(kwargs.get("month", "Jan"), 1)
        df = df[df["Trade Date"].dt.month == month_num]

    elif date_filter == "Custom Range":
        start_date = kwargs.get("start_date")
        end_date = kwargs.get("end_date")
        if start_date and end_date:
            df = df[
                (df["Trade Date"].dt.date >= start_date) &
                (df["Trade Date"].dt.date <= end_date)
            ]

    return df.to_dict("records")


def ensure_client_dirs(client_name: str) -> Path:
    base = Path("Clients") / client_name
    for sub in ["Processed_Reports", "Individual_Files", "Holdings", "temp_processing"]:
        (base / sub).mkdir(parents=True, exist_ok=True)
    return base


# ---------------------------------------------------------------------------
# Excel builders
# ---------------------------------------------------------------------------

def build_master_excel(
    trades: list,
    master_obligations: list,
    temp_path: Path,
    generate_tax: bool,
    holdings_file,
    corporate_actions,
    holdings_from_statement: dict | None,
) -> Path | None:
    try:
        master_df = pd.DataFrame(trades)

        # Ensure 'date' column exists for tax engine
        if "Trade Date" in master_df.columns:
            master_df["date"] = pd.to_datetime(master_df["Trade Date"], errors="coerce")
        elif "date" not in master_df.columns:
            master_df["date"] = None

        REQUIRED_COLS = [
            "Contract Note No", "Trade Date", "ISIN", "Security Name / Symbol",
            "Quantity (Buy)", "WAP (Across Exchanges) (Buy)", "Brokerage Per Share (Rs) (Buy)",
            "WAP (Across Exchanges) After Brokerage (Rs) (Buy)", "Total BUY Value After Brokerage",
            "Quantity (Sell)", "WAP (Across Exchanges) (Sell)", "Brokerage Per Share (Rs) (Sell)",
            "WAP (Across Exchanges) After Brokerage (Rs) (Sell)", "Total SELL Value After Brokerage",
            "Net Quantity", "Net Obligation For ISIN", "Net Settlement (Receivable/Payable)",
        ]
        NUMERIC_COLS = REQUIRED_COLS[4:]

        for col in REQUIRED_COLS:
            if col not in master_df.columns:
                master_df[col] = 0.0
        master_df = master_df[REQUIRED_COLS]

        for col in NUMERIC_COLS:
            master_df[col] = pd.to_numeric(master_df[col], errors="coerce").fillna(0.0)

        master_df = master_df.sort_values("Trade Date")

        # Grand total row
        grand_total = {c: "" for c in REQUIRED_COLS}
        grand_total.update({
            "ISIN": "TOTAL",
            "Quantity (Buy)": master_df["Quantity (Buy)"].sum(),
            "Quantity (Sell)": master_df["Quantity (Sell)"].sum(),
            "Net Obligation For ISIN": master_df["Net Obligation For ISIN"].sum(),
            "Net Settlement (Receivable/Payable)": master_df["Net Settlement (Receivable/Payable)"].sum(),
        })
        final_df = pd.concat([master_df, pd.DataFrame([grand_total])], ignore_index=True)

        # Output file
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = temp_path / f"CONSOLIDATED_TRADES_{ts}.xlsx"

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            # Sheet 1: Master Trades
            final_df.to_excel(writer, sheet_name="Master_Trades", index=False)
            try:
                from openpyxl.utils import get_column_letter
                ws = writer.sheets["Master_Trades"]
                col_letter = get_column_letter(17)
                for r in range(2, len(final_df) + 2):
                    ws[f"{col_letter}{r}"].number_format = "#,##0.00"
            except Exception:
                pass

            # Sheet 2: Master Tax Summary
            _write_tax_summary(writer, master_df)

            # Sheet 3: Master Obligations
            _write_obligations(writer, master_obligations)

            # Sheet 4: Capital Gains (optional)
            if generate_tax:
                enhanced_holdings = holdings_file
                if holdings_from_statement:
                    rows = []
                    for (broker, sec_name), lots in holdings_from_statement.items():
                        for lot in lots:
                            rows.append({
                                "ISIN": sec_name,
                                "Quantity": lot["quantity"],
                                "Purchase Date": lot["date"].strftime("%d-%m-%Y") if lot["date"] else "",
                                "Purchase Price": lot["rate"],
                                "Broker": broker,
                            })
                    if rows:
                        buf = io.StringIO()
                        pd.DataFrame(rows).to_csv(buf, index=False)
                        buf.seek(0)
                        enhanced_holdings = buf

                ok = create_capital_gains_summary_sheet(writer, final_df, enhanced_holdings, corporate_actions)
                if ok:
                    st.success("Tax Report Ready: Capital gains calculated using FIFO method")
                else:
                    st.warning("Tax calculation completed but no capital gains data found")

        return out_path

    except Exception as e:
        st.error(f"Error creating Excel: {e}")
        return None


def _write_tax_summary(writer, master_df: pd.DataFrame):
    try:
        pd.DataFrame({
            "Category": ["Total Buy Quantity", "Total Sell Quantity", "Total Net Obligation"],
            "Amount": [
                master_df["Quantity (Buy)"].sum(),
                master_df["Quantity (Sell)"].sum(),
                master_df["Net Obligation For ISIN"].sum(),
            ],
        }).to_excel(writer, sheet_name="Master_Tax_Summary", index=False)
    except Exception as e:
        print(f"[TAX-SUMMARY-ERROR] {e}")


def _write_obligations(writer, master_obligations: list):
    try:
        if not master_obligations:
            pd.DataFrame(columns=["Contract Note No", "Trade Date"]).to_excel(
                writer, sheet_name="Master_Obligations", index=False
            )
            return

        rows = []
        for obl in master_obligations:
            cn = obl.get("Contract Note No", "")
            td = obl.get("Trade Date", "")
            all_rows = obl.get("all_rows", [])
            raw_text = obl.get("raw_text", "")

            if all_rows:
                for row_data in all_rows:
                    rows.append({"Contract Note No": cn, "Trade Date": td, **row_data})
            elif raw_text:
                rows.append({
                    "Contract Note No": cn, "Trade Date": td,
                    "Raw Text": raw_text,
                    "Extraction Method": obl.get("extraction_method", "N/A"),
                })
            else:
                rows.append({
                    "Contract Note No": cn, "Trade Date": td,
                    "Net Settlement": obl.get("net_settlement", 0.0),
                    "Note": "Table structure not found",
                })

        pd.DataFrame(rows).to_excel(writer, sheet_name="Master_Obligations", index=False)
    except Exception as e:
        print(f"[OBLIGATIONS-SHEET-ERROR] {e}")


def build_individual_zip(individual_files: list, temp_path: Path) -> bytes | None:
    try:
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for fd in individual_files:
                path = create_excel_output(fd["trades"], fd["filename"])
                if path and os.path.exists(path):
                    zf.write(path, os.path.basename(path))
        buf.seek(0)
        return buf.getvalue()
    except Exception as e:
        st.error(f"Error creating ZIP: {e}")
        return None


# ---------------------------------------------------------------------------
# Core processing
# ---------------------------------------------------------------------------

def process_files(
    uploaded_files,
    pdf_password: str | None,
    create_individual: bool,
    show_preview: bool,
    generate_tax: bool,
    holdings_file,
    corporate_actions,
    holding_statement_file,
    date_filter: str,
    date_filter_kwargs: dict,
    active_client: str,
):
    try:
        st.cache_data.clear()
    except Exception:
        pass

    # Parse holding statement
    holdings_from_statement = None
    if holding_statement_file:
        holdings_from_statement = parse_holding_statement_balances(holding_statement_file)
        if holdings_from_statement:
            st.success(f"Loaded holding statement with {len(holdings_from_statement)} securities")
        else:
            st.warning("Could not parse holding statement — proceeding without cross-validation")

    # Temp folder
    temp_path = Path("Clients") / active_client / "temp_processing"
    temp_path.mkdir(parents=True, exist_ok=True)

    # Save uploads to disk
    temp_files = []
    for uf in uploaded_files:
        dest = temp_path / uf.name
        with open(dest, "wb") as f:
            f.write(uf.getbuffer())
        temp_files.append(str(dest))

    all_trades: list = []
    master_obligations: list = []
    individual_files_data: list = []
    success_count = 0
    failed_files: list = []

    # 4-step progress
    with st.status("Processing Contract Notes…", expanded=True) as status:
        status.write("🔓 Step 1: Decrypting PDFs…")
        st.progress(0.25)
        time.sleep(0.3)

        status.write("📊 Step 2: Extracting Trades…")
        st.progress(0.5)

        for pdf_path in temp_files:
            try:
                trades, obligation_data = extract_from_angel_one_pdf(
                    pdf_path, pdf_password or None
                )

                if not trades:
                    failed_files.append(os.path.basename(pdf_path))
                    continue

                # Fill Column 17 uniformly for all trades from this PDF
                if obligation_data and obligation_data.get("net_settlement") is not None:
                    net_settlement = obligation_data["net_settlement"]
                else:
                    net_settlement = 0.0
                    st.warning(
                        f"⚠ Net Settlement defaulted to 0.00 for {os.path.basename(pdf_path)}"
                    )

                for trade in trades:
                    trade["Net Settlement (Receivable/Payable)"] = net_settlement

                if obligation_data:
                    master_obligations.append(obligation_data)

                for trade in trades:
                    if trade.get("ISIN") != "TOTAL":
                        all_trades.append(trade)

                if create_individual:
                    individual_files_data.append({
                        "filename": os.path.basename(pdf_path),
                        "trades": trades,
                    })

                success_count += 1

            except Exception as e:
                failed_files.append(os.path.basename(pdf_path))
                st.error(f"Error processing {os.path.basename(pdf_path)}: {e}")

        status.write("💰 Step 3: Parsing Obligations…")
        st.progress(0.75)
        time.sleep(0.3)

        status.write("✅ Step 4: Final Reconciliation…")
        st.progress(1.0)
        time.sleep(0.3)

        status.update(label="Processing Complete!", state="complete", expanded=False)

    # Apply date filter
    all_trades = apply_date_filter(all_trades, date_filter, **date_filter_kwargs)

    if not all_trades:
        st.error("No trades were extracted. Please check the uploaded files.")
        return

    # Build master Excel
    excel_path = build_master_excel(
        all_trades, master_obligations, temp_path,
        generate_tax, holdings_file, corporate_actions, holdings_from_statement,
    )

    if not excel_path or not os.path.exists(excel_path):
        st.error("Master Excel generation failed.")
        return

    # Copy to client's Processed_Reports
    client_reports = Path("Clients") / active_client / "Processed_Reports"
    client_reports.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    shutil.copy(excel_path, client_reports / f"Report_{ts}.xlsx")

    with open(excel_path, "rb") as f:
        excel_data = f.read()

    # Persist to session state
    st.session_state.master_trades = all_trades
    st.session_state.master_obligations = master_obligations
    st.session_state.individual_files = individual_files_data
    st.session_state.processing_complete = True
    st.session_state.master_excel_data = excel_data
    st.session_state.master_excel_filename = f"CONSOLIDATED_TRADES_{ts}.xlsx"

    # Summary
    st.markdown('<div class="success-message">✅ Processing Completed Successfully!</div>', unsafe_allow_html=True)
    st.markdown("### 📊 Processing Summary")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total Files", len(temp_files))
    c2.metric("Processed", success_count)
    c3.metric("Failed", len(failed_files))
    c4.metric("Total Trades", len(all_trades))

    if failed_files:
        st.warning(f"Failed: {', '.join(failed_files)}")

    # Download buttons
    st.markdown("### 📥 Download Results")
    col_dl1, col_dl2 = st.columns([2, 1])

    with col_dl1:
        st.download_button(
            label="📥 Download Consolidated Excel",
            data=st.session_state.master_excel_data,
            file_name=st.session_state.master_excel_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )

    with col_dl2:
        if create_individual and individual_files_data:
            zip_data = build_individual_zip(individual_files_data, temp_path)
            if zip_data:
                st.download_button(
                    label="📦 Download Individual Files",
                    data=zip_data,
                    file_name=f"individual_files_{ts}.zip",
                    mime="application/zip",
                    use_container_width=True,
                )

    if show_preview and all_trades:
        st.markdown("### 👁 Data Preview")
        st.dataframe(pd.DataFrame(all_trades).head(50), use_container_width=True)

    st.balloons()


# ---------------------------------------------------------------------------
# Main UI
# ---------------------------------------------------------------------------

def main():
    st.markdown('<h1 class="main-header">MRACA Smart Contract Note Converter</h1>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">Professional Financial Auditor v1.0.0</p>', unsafe_allow_html=True)

    # Sidebar
    with st.sidebar:
        st.markdown("### 📊 System Status")
        st.markdown("---")
        col1, col2 = st.columns(2)
        col1.metric("Total PDFs", st.session_state.total_pdfs_uploaded)
        col2.metric("Client Active", "✓" if st.session_state.active_client else "✗")

        if st.button("🔄 Reset Environment", use_container_width=True):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()

        st.markdown("---")
        if st.session_state.active_client:
            st.success(f"📁 {st.session_state.active_client}")
        else:
            st.info("No client selected")

    # ---- PHASE 1: Client selection ----
    if not st.session_state.active_client:
        st.markdown('<h2 class="client-greeting">👋 Welcome, MR! How can I help you today?</h2>', unsafe_allow_html=True)

        col1, col2 = st.columns(2)
        if col1.button("➕ Create New Client", use_container_width=True, type="primary"):
            st.session_state.show_client_creation = True
            st.session_state.show_client_selection = False
            st.rerun()
        if col2.button("📂 Select Existing Client", use_container_width=True):
            st.session_state.show_client_selection = True
            st.session_state.show_client_creation = False
            st.rerun()

        if st.session_state.show_client_creation:
            st.markdown("### ➕ Create New Client")
            st.markdown("---")
            client_name = st.text_input("Client Name", placeholder="Enter client name…")
            c1, c2 = st.columns(2)
            if c1.button("Create Client", type="primary"):
                if client_name and client_name.strip():
                    ensure_client_dirs(client_name.strip())
                    st.session_state.active_client = client_name.strip()
                    st.session_state.show_client_creation = False
                    st.rerun()
                else:
                    st.error("Please enter a client name")
            if c2.button("Cancel"):
                st.session_state.show_client_creation = False
                st.rerun()

        if st.session_state.show_client_selection:
            st.markdown("### 📂 Select Existing Client")
            st.markdown("---")
            clients_dir = Path("Clients")
            existing = [d.name for d in clients_dir.iterdir() if d.is_dir()] if clients_dir.exists() else []
            if existing:
                selected = st.selectbox("Select Client", existing)
                c1, c2 = st.columns(2)
                if c1.button("Select", type="primary"):
                    st.session_state.active_client = selected
                    st.session_state.show_client_selection = False
                    st.rerun()
                if c2.button("Cancel"):
                    st.session_state.show_client_selection = False
                    st.rerun()
            else:
                st.info("No existing clients found. Create a new client first.")
                if st.button("Cancel"):
                    st.session_state.show_client_selection = False
                    st.rerun()
        return  # stop here until a client is chosen

    # ---- PHASE 2–5: Guided workflow ----

    # Broker selection
    st.markdown("### 📋 Broker Selection")
    st.markdown("---")
    st.selectbox("Select Broker", ["Angel One", "Kotak", "Axis"],
                 help="Select the broker for the contract notes you are uploading")

    # Upload hub
    st.markdown("### 📤 Upload Hub")
    st.markdown("---")
    uploaded_files = st.file_uploader(
        "Drag & drop PDF files here or click to browse",
        type=["pdf"],
        accept_multiple_files=True,
        key="pdf_uploader",
    )
    if uploaded_files:
        st.success(f"✓ {len(uploaded_files)} file(s) uploaded")
        st.session_state.total_pdfs_uploaded = len(uploaded_files)

    password_protected = st.checkbox("PDFs are password protected")
    pdf_password = st.text_input("Enter PDF Password", type="password") if password_protected else None

    # Customisation & date filtering
    st.markdown("### ⚙️ Customisation & Date Filtering")
    st.markdown("---")

    col_left, col_right = st.columns(2)

    with col_left:
        report_options = st.multiselect(
            "Report Options",
            ["Generate Individual Files", "Show Data Preview", "Generate FY 2025-26 Tax Summary"],
            default=["Show Data Preview"],
        )

        date_filter = st.selectbox(
            "Date Period Filter",
            ["All Data", "Quarterly (Q1-Q4)", "Monthly", "Custom Range"],
        )

        # Date filter sub-controls (always define variables with safe defaults)
        quarter = None
        month = None
        start_date = None
        end_date = None

        if date_filter == "Quarterly (Q1-Q4)":
            quarter = st.selectbox("Select Quarter", list(QUARTER_MAP.keys()))
        elif date_filter == "Monthly":
            month = st.selectbox("Select Month", list(MONTH_MAP.keys()))
        elif date_filter == "Custom Range":
            cd1, cd2 = st.columns(2)
            start_date = cd1.date_input("Start Date")
            end_date = cd2.date_input("End Date")

    with col_right:
        # Advanced modules
        st.markdown("### 📦 Advanced Modules")
        st.markdown("---")

        holding_statement_file = None
        with st.expander("📑 Holding Statement (Excel)"):
            holding_statement_file = st.file_uploader(
                "Upload Holding Statement Excel",
                type=["xlsx", "xls"],
                help="Excel file with 'Balances' sheet for cross-validation",
            )
            if holding_statement_file:
                st.success("✓ Holding statement loaded")

        holdings_file = None
        with st.expander("📦 Opening Holdings (Optional)"):
            holdings_file = st.file_uploader(
                "Upload Opening Holdings CSV",
                type=["csv"],
                help="Columns: ISIN, Quantity, Purchase Date, Purchase Price",
            )

        with st.expander("🏢 Corporate Actions"):
            corporate_actions = st.data_editor(
                pd.DataFrame(columns=["ISIN", "Action Type", "Ratio", "Effective Date"]),
                column_config={
                    "ISIN": st.column_config.TextColumn("ISIN"),
                    "Action Type": st.column_config.SelectboxColumn(
                        "Action Type", options=["", "Bonus", "Split"]
                    ),
                    "Ratio": st.column_config.TextColumn("Ratio"),
                    "Effective Date": st.column_config.DateColumn("Effective Date"),
                },
                num_rows="dynamic",
                use_container_width=True,
            )

    # Process button
    if uploaded_files and st.button(
        "🚀 Process Contract Notes", type="primary", use_container_width=True
    ):
        date_filter_kwargs = {
            "quarter": quarter,
            "month": month,
            "start_date": start_date,
            "end_date": end_date,
        }
        process_files(
            uploaded_files=uploaded_files,
            pdf_password=pdf_password,
            create_individual="Generate Individual Files" in report_options,
            show_preview="Show Data Preview" in report_options,
            generate_tax="Generate FY 2025-26 Tax Summary" in report_options,
            holdings_file=holdings_file,
            corporate_actions=corporate_actions,
            holding_statement_file=holding_statement_file,
            date_filter=date_filter,
            date_filter_kwargs=date_filter_kwargs,
            active_client=st.session_state.active_client,
        )

    # ---- PHASE 5: Report Vault ----
    if st.session_state.processing_complete:
        st.markdown("### 📊 Report Vault")
        st.markdown("---")

        tab1, tab2, tab3 = st.tabs(["📊 Master Trades", "🧾 Obligation Breakdown", "📈 Summary"])

        with tab1:
            if st.session_state.master_trades:
                st.dataframe(pd.DataFrame(st.session_state.master_trades), use_container_width=True)
            else:
                st.info("No trade data available")

        with tab2:
            if st.session_state.master_obligations:
                st.dataframe(pd.DataFrame(st.session_state.master_obligations), use_container_width=True)
            else:
                st.info("No obligation data available")

        with tab3:
            if st.session_state.master_trades:
                df_t = pd.DataFrame(st.session_state.master_trades)
                c1, c2, c3 = st.columns(3)
                c1.metric("Total Buy",  df_t["Quantity (Buy)"].sum())
                c2.metric("Total Sell", df_t["Quantity (Sell)"].sum())
                c3.metric("Net Settlement", df_t["Net Settlement (Receivable/Payable)"].sum())

        # Historical reports
        st.markdown("### 📁 Historical Reports")
        st.markdown("---")
        reports_path = Path("Clients") / st.session_state.active_client / "Processed_Reports"
        report_files = sorted(reports_path.glob("*.xlsx"), reverse=True) if reports_path.exists() else []

        if report_files:
            for report_file in report_files:
                rc1, rc2 = st.columns([4, 1])
                rc1.markdown(f"📄 {report_file.name}")
                with open(report_file, "rb") as f:
                    rc2.download_button(
                        "Download",
                        f.read(),
                        report_file.name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=report_file.name,
                    )
        else:
            st.info("No historical reports found")


if __name__ == "__main__":
    main()
